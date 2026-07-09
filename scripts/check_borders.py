"""Check if cell borders are preserved in the template."""
import openpyxl
from openpyxl.styles import Border

print("=" * 80)
print("ORIGINAL TRANSIMITALS.xlsx - Sheet '170' - Cell Borders in item area")
print("=" * 80)
wb = openpyxl.load_workbook('/home/z/my-project/upload/TRANSIMITALS.xlsx')
ws = wb['170']

for r in range(14, 22):
    for c in range(2, 9):
        cell = ws.cell(row=r, column=c)
        b = cell.border
        sides = []
        if b.left and b.left.style: sides.append(f'L={b.left.style}')
        if b.right and b.right.style: sides.append(f'R={b.right.style}')
        if b.top and b.top.style: sides.append(f'T={b.top.style}')
        if b.bottom and b.bottom.style: sides.append(f'B={b.bottom.style}')
        if sides:
            print(f'  R{r} C{c}: {",".join(sides)} | fill={cell.fill.fgColor.rgb if cell.fill and cell.fill.patternType else "none"}')

print()
print("=" * 80)
print("CURRENT TEMPLATE - Cell Borders in item area")
print("=" * 80)
wb2 = openpyxl.load_workbook('/home/z/my-project/public/templates/TRANSIMITALS_template.xlsx')
ws2 = wb2.active

for r in range(14, 22):
    for c in range(2, 9):
        cell = ws2.cell(row=r, column=c)
        b = cell.border
        sides = []
        if b.left and b.left.style: sides.append(f'L={b.left.style}')
        if b.right and b.right.style: sides.append(f'R={b.right.style}')
        if b.top and b.top.style: sides.append(f'T={b.top.style}')
        if b.bottom and b.bottom.style: sides.append(f'B={b.bottom.style}')
        if sides:
            print(f'  R{r} C{c}: {",".join(sides)} | fill={cell.fill.fgColor.rgb if cell.fill and cell.fill.patternType else "none"}')

print()
print("=" * 80)
print("Column widths comparison")
print("=" * 80)
print("Original sheet '170':")
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    cd = ws.column_dimensions.get(col_letter)
    print(f'  {col_letter}: width={cd.width if cd else "default"}')
print("Current template:")
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    cd = ws2.column_dimensions.get(col_letter)
    print(f'  {col_letter}: width={cd.width if cd else "default"}')

print()
print("=" * 80)
print("Row heights comparison")
print("=" * 80)
print("Original sheet '170':")
for r in range(1, 35):
    rd = ws.row_dimensions.get(r)
    if rd and rd.height:
        print(f'  R{r}: height={rd.height}')
print("Current template:")
for r in range(1, 35):
    rd = ws2.row_dimensions.get(r)
    if rd and rd.height:
        print(f'  R{r}: height={rd.height}')
