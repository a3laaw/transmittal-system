import openpyxl
import re

wb = openpyxl.load_workbook('/home/z/my-project/upload/TRANSIMITALS.xlsx', data_only=True)

# Check header row (row 14) for all sheets with Rev.01 or Rev.02
included_sheets = []
for sn in wb.sheetnames:
    ws = wb[sn]
    rev_str = str(ws.cell(row=3, column=9).value or '').strip()
    rev_match = re.search(r'Rev\.?\s*0?(\d+)', rev_str, re.IGNORECASE)
    rev_num = rev_match.group(1) if rev_match else '0'
    if rev_num != '0':
        included_sheets.append((sn, rev_num))

print(f'Total sheets with Rev.01 or Rev.02: {len(included_sheets)}')
print()

# Look at row 14 and 15 for each
print('Header inspection for included sheets:')
print('=' * 100)
for sn, rev_num in included_sheets:
    ws = wb[sn]
    print(f'\n--- Sheet: {sn!r} (Rev.{rev_num}) ---')
    for r in [14, 15]:
        row_vals = []
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            row_vals.append(f'C{c}={str(v)!r}' if v else f'C{c}=--')
        print(f'  R{r}: ' + ' | '.join(row_vals))
