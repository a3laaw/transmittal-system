"""
Generate a new Excel transmittal file based on the template.

Usage:
    python3 gen_excel_template.py <template_path> <output_path> <reference> <date_YYYY-MM-DD> <description>

The template is TRANSIMITALS_template.xlsx.
This script:
1. Loads the template (which has all borders/formatting intact)
2. Sets Transmittal No (G3), Rev.00 (I3), Date (G4)
3. Fills the description into the first item row (row 16)
4. Sets QTY=1 in the first item row to make the table look "active"
5. Saves as a new file with the reference as sheet name
"""
import openpyxl
import sys
import re
from datetime import datetime

TEMPLATE = sys.argv[1]
OUT = sys.argv[2]
REFERENCE = sys.argv[3]
DATE_STR = sys.argv[4]
DESCRIPTION = sys.argv[5]

def main():
    print(f'Loading template: {TEMPLATE}', flush=True)
    wb = openpyxl.load_workbook(TEMPLATE)

    # Use the first sheet
    ws = wb[wb.sheetnames[0]]
    print(f'Using sheet: {ws.title}', flush=True)

    # Build Transmittal No string
    transmittal_no_str = f'Transmittal No:{REFERENCE}'
    rev_str = 'Rev.00'

    # Parse date
    try:
        dt = datetime.strptime(DATE_STR, '%Y-%m-%d')
        date_display = dt.strftime('%d/%m/%Y')
    except Exception:
        date_display = DATE_STR

    date_str = f'Date :{date_display}'

    print(f'  Setting G3 = {transmittal_no_str!r}', flush=True)
    print(f'  Setting I3 = {rev_str!r}', flush=True)
    print(f'  Setting G4 = {date_str!r}', flush=True)

    # Set header cells
    ws.cell(row=3, column=7).value = transmittal_no_str  # G3
    ws.cell(row=3, column=9).value = rev_str              # I3
    ws.cell(row=4, column=7).value = date_str             # G4

    # Fill description into the first item row (row 16)
    # The DESCRIPTION column is in E (5) for most sheets, F (6) for some
    # Looking at original sheet 170: description was in C5 (E=5)
    # Let's set both E and F to be safe — but actually F is the description column in 170
    # Wait, looking at row 14 of sheet 170:
    #   C2='QTY' | C3='DRWS. SPEC. O' | C4='ITEM SEQ' | C5='DESCRIPTION' | C7='(+)TYP'
    # So DESCRIPTION is in column E (5).
    # But in sheet 27 it's in column F (6):
    #   C2='QTY' | C3='DRWS. SPEC. O' | C4='ITEM SEQ' | C6='DESCRIPTION' | C7='(+)TYP'
    # We need to detect the right column.

    desc_col = None
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=14, column=c).value or '').upper()
        if 'DESCRIPTION' in v:
            desc_col = c
            break
    if desc_col is None:
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(row=15, column=c).value or '').upper()
            if 'DESCRIPTION' in v:
                desc_col = c
                break
    if desc_col is None:
        # Default to E (5)
        desc_col = 5

    print(f'  DESCRIPTION column detected: {desc_col}', flush=True)

    # Fill description into item rows (starting from row 16)
    # Each line in the description (split by &, /, \n, or comma) goes in a SEPARATE row
    if DESCRIPTION:
        # Split description into individual lines
        # Common separators in the data: &, /, newline, comma
        import re as re2
        parts = re2.split(r'\s*[&/]\s*|\n|,(?=\s)', DESCRIPTION.strip())
        parts = [p.strip() for p in parts if p.strip()]

        if not parts:
            parts = [DESCRIPTION]

        print(f'  Description split into {len(parts)} line(s):', flush=True)
        for i, p in enumerate(parts):
            print(f'    {i+1}. {p!r}', flush=True)

        # Fill each part into a separate row (row 16, 17, 18, ...)
        # Item rows go from 16 to 26 (11 rows available)
        max_rows = 11  # rows 16-26
        for i, part in enumerate(parts[:max_rows]):
            row_num = 16 + i
            # Only fill DESCRIPTION - leave QTY, DRWS SPEC, ITEM SEQ, (+)TYP empty
            # for the user to fill manually after downloading

            # Set description
            desc_cell = ws.cell(row=row_num, column=desc_col)
            if not isinstance(desc_cell, openpyxl.cell.cell.MergedCell):
                desc_cell.value = part
                print(f'  Set R{row_num}C{desc_col} (description) = {part!r}', flush=True)

        if len(parts) > max_rows:
            print(f'  WARNING: {len(parts) - max_rows} line(s) truncated (only {max_rows} rows available)', flush=True)
    else:
        # No description - leave all columns empty for user to fill
        pass

    # Rename sheet to reference (strip non-alphanumeric, max 31 chars)
    sheet_name = re.sub(r'[^\w\-]', '', REFERENCE)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    ws.title = sheet_name
    print(f'  Renamed sheet to: {sheet_name!r}', flush=True)

    print(f'Saving to: {OUT}', flush=True)
    wb.save(OUT)
    print('Done.', flush=True)

    # Verify
    print()
    print('Verification:', flush=True)
    wb2 = openpyxl.load_workbook(OUT)
    ws2 = wb2[wb2.sheetnames[0]]
    print(f'  Sheet: {ws2.title}')
    print(f'  G3 (Transmittal No): {ws2.cell(row=3, column=7).value}')
    print(f'  I3 (Rev): {ws2.cell(row=3, column=9).value}')
    print(f'  G4 (Date): {ws2.cell(row=4, column=7).value}')
    print(f'  B16 (QTY): {ws2.cell(row=16, column=2).value}')
    print(f'  R16C{desc_col} (Description): {ws2.cell(row=16, column=desc_col).value}')

    # Verify borders are preserved
    bordered_count = 0
    for r in range(14, 27):
        for c in range(2, 9):
            cell = ws2.cell(row=r, column=c)
            b = cell.border
            if (b.left and b.left.style) or (b.right and b.right.style) or (b.top and b.top.style) or (b.bottom and b.bottom.style):
                bordered_count += 1
    print(f'  Bordered cells in table area: {bordered_count}')

if __name__ == '__main__':
    main()
