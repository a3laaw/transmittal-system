import openpyxl
wb = openpyxl.load_workbook('/home/z/my-project/upload/TRANSIMITALS.xlsx', data_only=True)

# Look at sheets with revisions (Rev.01, Rev.02, etc.)
for sn in ['1r1', '27-L01' if '27-L01' in wb.sheetnames else '27', '3r1', '3r2', '18R1', '46-L01' if '46-L01' in wb.sheetnames else '46']:
    # Find actual revision sheets
    pass

# Find sheets that look like revisions
import re
rev_sheets = []
for sn in wb.sheetnames:
    # Look for patterns like "1r1", "27-R01", "18R1", "59 R01", etc.
    if re.search(r'(?i)r\s*0?\d', sn) or re.search(r'(?i)-r\d', sn) or re.search(r'(?i)r\d', sn):
        rev_sheets.append(sn)

print('Sheets that look like revisions:')
print(rev_sheets[:30])
print('Total:', len(rev_sheets))
print()

# Look at a few of these in detail
for sn in rev_sheets[:5]:
    print('=' * 80)
    print('SHEET:', sn)
    print('=' * 80)
    ws = wb[sn]
    for r in range(1, min(ws.max_row + 1, 30)):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row_vals.append('')
            else:
                row_vals.append(str(v).replace('\n', ' / '))
        # Only print non-empty rows
        if any(x.strip() for x in row_vals):
            print(f"R{r:02d}: " + ' | '.join(row_vals))
    print()
