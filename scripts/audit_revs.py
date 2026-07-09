import openpyxl
import re
from collections import Counter

wb = openpyxl.load_workbook('/home/z/my-project/upload/TRANSIMITALS.xlsx', data_only=True)

rev_counter = Counter()
sheets_with_no_rev = []
all_rev_info = []

for sn in wb.sheetnames:
    ws = wb[sn]
    # Rev is in column I (9), row 3
    rev_cell = ws.cell(row=3, column=9).value
    transmittal_cell = ws.cell(row=3, column=7).value
    date_cell = ws.cell(row=4, column=7).value
    
    rev_str = str(rev_cell).strip() if rev_cell else ''
    transmittal_str = str(transmittal_cell).strip() if transmittal_cell else ''
    date_str = str(date_cell).strip() if date_cell else ''
    
    # Extract Rev number
    rev_match = re.search(r'Rev\.?\s*0?(\d+)', rev_str, re.IGNORECASE)
    rev_num = rev_match.group(1) if rev_match else None
    
    # Extract Transmittal No
    tn_match = re.search(r'Transmittal\s*No:?\s*([A-Z\-]*\s*\d+)', transmittal_str, re.IGNORECASE)
    tn = tn_match.group(1).strip() if tn_match else None
    
    # Extract Date
    date_match = re.search(r'Date\s*:?\s*(\S+)', date_str, re.IGNORECASE)
    date_val = date_match.group(1).strip() if date_match else None
    
    rev_counter[rev_str or 'EMPTY'] += 1
    if not rev_str:
        sheets_with_no_rev.append(sn)
    
    all_rev_info.append({
        'sheet': sn,
        'rev_raw': rev_str,
        'rev_num': rev_num,
        'transmittal_raw': transmittal_str,
        'transmittal_no': tn,
        'date_raw': date_str,
        'date_val': date_val
    })

print('Rev value distribution:')
for k, v in sorted(rev_counter.items()):
    print(f'  {k!r}: {v} sheets')
print()
print(f'Sheets with no Rev: {len(sheets_with_no_rev)}')
for s in sheets_with_no_rev:
    print(f'  - {s}')
print()
print('Sample of all_rev_info:')
for r in all_rev_info[:10]:
    print(r)
