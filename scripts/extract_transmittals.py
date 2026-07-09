"""
Extract Transmittal data from TRANSIMITALS.xlsx
- Include only sheets where Rev is 01, 02, 03, ... (NOT Rev.00 / R00)
- Extract: Transmittal No, Rev, Date, DESCRIPTION, (+)TYP
- Output: a new Excel file with one row per DESCRIPTION line
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import os

SRC = '/home/z/my-project/upload/TRANSIMITALS.xlsx'
DST = '/home/z/my-project/download/TRANSMITTALS_Extracted_Rev01plus.xlsx'

# ---------- Helpers ----------
def clean(v):
    """Convert cell value to a clean string, replacing newlines with ' / '."""
    if v is None:
        return ''
    s = str(v).strip()
    # Replace newlines with ' / ' for readability
    s = re.sub(r'\s*\n\s*', ' / ', s)
    # Collapse multiple spaces
    s = re.sub(r' {2,}', ' ', s).strip()
    return s

def parse_transmittal_no(raw):
    """Parse 'Transmittal No: CI-001' -> 'CI-001'"""
    if not raw:
        return ''
    m = re.search(r'Transmittal\s*No:?\s*([A-Za-z\-]*\s*\d+)', raw, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    # Fallback: take everything after the colon
    if ':' in raw:
        return raw.split(':', 1)[1].strip()
    return raw.strip()

def parse_rev(raw):
    """Parse 'Rev.01' -> '01'"""
    if not raw:
        return ''
    m = re.search(r'Rev\.?\s*0?(\d+)', raw, re.IGNORECASE)
    if m:
        return m.group(1).zfill(2)  # '01', '02', etc.
    return raw.strip()

def parse_date(raw):
    """Parse 'Date :24/06/2026' -> '24/06/2026'"""
    if not raw:
        return ''
    m = re.search(r'Date\s*:?\s*(\S+)', raw, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return raw.strip()

def find_header_col(ws, header_row, header_text):
    """Find the column index where header_text appears in header_row."""
    for c in range(1, ws.max_column + 1):
        v = clean(ws.cell(row=header_row, column=c).value)
        if v and header_text.lower() in v.lower():
            return c
    return None

# ---------- Main extraction ----------
print(f'Loading source: {SRC}')
wb_src = openpyxl.load_workbook(SRC, data_only=True)
print(f'Total sheets in source: {len(wb_src.sheetnames)}')

rows_out = []  # list of dicts
stats = {
    'total_sheets': 0,
    'rev00_skipped': 0,
    'rev01plus_included': 0,
    'rows_extracted': 0,
}

# Sort sheet names by natural order (so 1, 2, 3, ... 10, 11, ... not 1, 10, 11, 2)
def natural_key(s):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]

sorted_sheets = sorted(wb_src.sheetnames, key=natural_key)

for sn in sorted_sheets:
    stats['total_sheets'] += 1
    ws = wb_src[sn]

    # Parse header cells (Row 3, Column G = Transmittal No; Row 3, Column I = Rev; Row 4, Column G = Date)
    transmittal_raw = ws.cell(row=3, column=7).value
    rev_raw = ws.cell(row=3, column=9).value
    date_raw = ws.cell(row=4, column=7).value

    transmittal_no = parse_transmittal_no(clean(transmittal_raw))
    rev_num = parse_rev(clean(rev_raw))  # '00', '01', '02', ...
    date_val = parse_date(clean(date_raw))

    # Filter: SKIP Rev.00 (R00)
    if rev_num == '00' or rev_num == '0':
        stats['rev00_skipped'] += 1
        continue

    stats['rev01plus_included'] += 1

    # Find DESCRIPTION and (+)TYP columns from header row 14
    desc_col = find_header_col(ws, 14, 'DESCRIPTION')
    typ_col = find_header_col(ws, 14, '(+)TYP')

    if desc_col is None:
        # Fallback: try row 15
        desc_col = find_header_col(ws, 15, 'DESCRIPTION')
    if typ_col is None:
        typ_col = find_header_col(ws, 15, '(+)TYP')

    # Default if still not found
    if desc_col is None:
        desc_col = 5  # column E
    if typ_col is None:
        typ_col = 7  # column G

    # Some sheets (e.g. 27-R01, 27-R02) have DESCRIPTION header in column F
    # but the actual data is in column E. To be safe, always consider BOTH
    # columns E (5) and F (6) as candidate DESCRIPTION columns.
    desc_candidate_cols = sorted(set([desc_col, 5, 6]))
    # (+)TYP is consistently in column G (7) across all sheets.
    typ_candidate_cols = sorted(set([typ_col, 7]))

    # Extract DESCRIPTION and (+)TYP entries from row 16 onward.
    # Stop conditions (boilerplate rows that come after the data table):
    #   - "SUBMITTAL IS AS PER SPECS" (any column)
    #   - "SUBSTITUE SUBMITTAL"        (any column)
    #   - "REMARKS:"                   (any column)
    #   - "FOR CONTRACTOR"             (any column)
    #   - "Engineer's Rep"             (any column)
    STOP_PATTERNS = [
        'SUBMITTAL IS AS PER SPECS',
        'SUBSTITUE SUBMITTAL',
        'REMARKS:',
        'FOR CONTRACTOR',
        "ENGINEER'S REP",
    ]

    items = []
    for r in range(16, ws.max_row + 1):
        # Collect all cell values in this row (columns 1..9) for stop-condition checking
        row_cells = [clean(ws.cell(row=r, column=c).value) for c in range(1, 10)]
        row_text_upper = ' | '.join(row_cells).upper()

        # Stop scanning at the first boilerplate row
        stop = False
        for pat in STOP_PATTERNS:
            if pat in row_text_upper:
                stop = True
                break
        if stop:
            break

        desc_val = ''
        for cc in desc_candidate_cols:
            v = clean(ws.cell(row=r, column=cc).value)
            if v:
                desc_val = v
                break
        typ_val = ''
        for cc in typ_candidate_cols:
            v = clean(ws.cell(row=r, column=cc).value)
            if v:
                typ_val = v
                break

        # Read context columns to decide if the row has real content
        qty_val       = clean(ws.cell(row=r, column=2).value)
        boq_val       = clean(ws.cell(row=r, column=3).value)
        item_seq_val  = clean(ws.cell(row=r, column=4).value)

        row_has_content = any([desc_val, typ_val, qty_val, boq_val, item_seq_val])
        if not row_has_content:
            continue

        items.append({
            'description': desc_val,
            'typ': typ_val,
            'row': r,
        })

    # If no items found, still create one row with empty DESCRIPTION/(+)TYP so the sheet is recorded
    if not items:
        items = [{'description': '', 'typ': '', 'row': 0}]

    for it in items:
        stats['rows_extracted'] += 1
        rows_out.append({
            'sheet_name': sn,
            'transmittal_no': transmittal_no,
            'rev': f'Rev.{rev_num}',
            'date': date_val,
            'description': it['description'],
            'typ': it['typ'],
        })

# ---------- Write output ----------
print(f'\nWriting output: {DST}')
os.makedirs(os.path.dirname(DST), exist_ok=True)

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = 'Extracted (Rev.01+)'

# Headers
headers = ['Sheet Name', 'Transmittal No', 'Rev', 'Date', 'DESCRIPTION', '(+)TYP']
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(border_style='thin', color='B0B0B0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, h in enumerate(headers, 1):
    cell = ws_out.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border

# Data rows
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
data_font = Font(name='Calibri', size=10)
data_align_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
data_align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Group rows by sheet so we can apply alternating colors per sheet
prev_sheet = None
color_toggle = False
for i, row in enumerate(rows_out, 2):
    if row['sheet_name'] != prev_sheet:
        color_toggle = not color_toggle
        prev_sheet = row['sheet_name']

    values = [
        row['sheet_name'],
        row['transmittal_no'],
        row['rev'],
        row['date'],
        row['description'],
        row['typ'],
    ]
    for c, v in enumerate(values, 1):
        cell = ws_out.cell(row=i, column=c, value=v)
        cell.font = data_font
        cell.border = border
        if c in (1, 2, 3, 4, 6):  # center columns
            cell.alignment = data_align_center
        else:
            cell.alignment = data_align_top
        if color_toggle:
            cell.fill = alt_fill

# Column widths
col_widths = {
    'A': 16,  # Sheet Name
    'B': 18,  # Transmittal No
    'C': 10,  # Rev
    'D': 14,  # Date
    'E': 60,  # DESCRIPTION
    'F': 12,  # (+)TYP
}
for col_letter, w in col_widths.items():
    ws_out.column_dimensions[col_letter].width = w

# Freeze the header row
ws_out.freeze_panes = 'A2'

# Auto-filter on the header
ws_out.auto_filter.ref = f'A1:F{len(rows_out) + 1}'

# Workbook metadata
wb_out.properties.creator = 'Z.ai'
wb_out.properties.title = 'Transmittals Extracted (Rev.01+)'

wb_out.save(DST)

print(f'\nDone. Output saved to: {DST}')
print(f'\nStats:')
for k, v in stats.items():
    print(f'  {k}: {v}')
print(f'\nFirst 10 rows of output:')
for r in rows_out[:10]:
    print(f'  {r}')
