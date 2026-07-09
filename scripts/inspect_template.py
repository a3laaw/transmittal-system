"""Inspect the current template and compare with original."""
import openpyxl

print("=" * 80)
print("CURRENT TEMPLATE")
print("=" * 80)
wb = openpyxl.load_workbook('/home/z/my-project/public/templates/TRANSIMITALS_template.xlsx')
ws = wb.active
print(f'Sheet: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}')
print(f'Merged ranges: {len(ws.merged_cells.ranges)}')
for r in range(1, 35):
    row_vals = []
    for c in range(1, 10):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            row_vals.append(f'C{c}={str(v)[:40]!r}')
    if row_vals:
        print(f'  R{r:02d}: ' + ' | '.join(row_vals))

print()
print("=" * 80)
print("ORIGINAL TEMPLATE (sheet '170' from TRANSIMITALS.xlsx)")
print("=" * 80)
wb2 = openpyxl.load_workbook('/home/z/my-project/upload/TRANSIMITALS.xlsx')
ws2 = wb2['170']
print(f'Sheet: {ws2.title}, max_row={ws2.max_row}, max_col={ws2.max_column}')
print(f'Merged ranges: {len(ws2.merged_cells.ranges)}')
for r in range(1, 35):
    row_vals = []
    for c in range(1, 10):
        v = ws2.cell(row=r, column=c).value
        if v is not None:
            row_vals.append(f'C{c}={str(v)[:40]!r}')
    if row_vals:
        print(f'  R{r:02d}: ' + ' | '.join(row_vals))
