"""
Combine all sheets from LOG_Final.xlsm into one single sheet.
- Source sheets: INPUT, CIV, EL, PL, HVAC, FF, ELVE, CHECK LIST, MOH, soor, CONT
  (00-Dashboard is excluded - it's a summary dashboard, not a register)
- Output: one unified table on a single sheet
- Each row prefixed with: Source Sheet + Discipline (derived)
- All columns preserved (30 max)
- Empty rows kept as-is
- Sorted by Reference
- Simple Excel without colors
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from datetime import datetime

SRC = '/home/z/my-project/upload/LOG_Final.xlsm'
DST = '/home/z/my-project/download/LOG_Combined_AllSheets.xlsx'

# ---------- Helpers ----------
def clean(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        # Format dates as DD/MM/YYYY
        return v.strftime('%d/%m/%Y')
    s = str(v).strip()
    # Replace newlines with ' / ' for readability
    s = re.sub(r'\s*\n\s*', ' / ', s)
    s = re.sub(r' {2,}', ' ', s).strip()
    return s

def is_empty_row(values):
    """Check if all values in a row are empty."""
    return all(v == '' for v in values)

def discipline_from_sheet(sheet_name):
    """Derive a Discipline code from the sheet name."""
    sn = sheet_name.strip().upper()
    mapping = {
        'INPUT': 'INPUT',
        'CIV': 'CIV',
        'EL': 'EL',
        'PL': 'PL',
        'HVAC': 'HVAC',
        'FF': 'FF',
        'ELVE': 'ELEV',
        'CHECK LIST': 'CHKLIST',
        'MOH': 'MOH',
        'SOOR': 'SOOR',
        'CONT': 'CONT',
    }
    return mapping.get(sn, sheet_name)

# ---------- Load source ----------
print(f'Loading source: {SRC}')
wb_src = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
print(f'Source sheets: {wb_src.sheetnames}')

# ---------- Define the unified header ----------
# All discipline sheets share this 30-column structure (CIV, EL, PL, HVAC, FF, ELVE):
#   # | type | Reference | Description | REV.0..REV.7 (3 cols each = 24) | Consultant | moh
# INPUT has 30 cols but with Discipline as col 2 instead of "type":
#   # | Discipline | Type | Reference | Description | REV.0..REV.7 (24) | Status
# CHECK LIST, MOH, soor, CONT have 29 cols (no Consultant/moh split — single Status col)
#
# Unified output header (32 cols):
#   Source Sheet | Discipline | # | Type | Reference | Description |
#   REV.0 Submit | REV.0 Date Reply | REV.0 Action |
#   REV.1 Submit | REV.1 Date Reply | REV.1 Action |
#   REV.2 Submit | REV.2 Date Reply | REV.2 Action |
#   REV.3 Submit | REV.3 Date Reply | REV.3 Action |
#   REV.4 Submit | REV.4 Date Reply | REV.4 Action |
#   REV.5 Submit | REV.5 Date Reply | REV.5 Action |
#   REV.6 Submit | REV.6 Date Reply | REV.6 Action |
#   REV.7 Submit | REV.7 Date Reply | REV.7 Action |
#   Consultant | MOH/Status

UNIFIED_HEADERS = [
    'Source Sheet', 'Discipline', '#', 'Type', 'Reference', 'Description',
    'REV.0 Submit', 'REV.0 Date Reply', 'REV.0 Action',
    'REV.1 Submit', 'REV.1 Date Reply', 'REV.1 Action',
    'REV.2 Submit', 'REV.2 Date Reply', 'REV.2 Action',
    'REV.3 Submit', 'REV.3 Date Reply', 'REV.3 Action',
    'REV.4 Submit', 'REV.4 Date Reply', 'REV.4 Action',
    'REV.5 Submit', 'REV.5 Date Reply', 'REV.5 Action',
    'REV.6 Submit', 'REV.6 Date Reply', 'REV.6 Action',
    'REV.7 Submit', 'REV.7 Date Reply', 'REV.7 Action',
    'Consultant', 'MOH/Status',
]
# 32 columns total (indices 0..31)

def map_input_row(row_values):
    """Map an INPUT sheet row (30 cols) to the unified 32-col output.
    INPUT layout: #(1) Discipline(2) Type(3) Reference(4) Description(5)
                  REV.0(6,7,8) REV.1(9,10,11) ... REV.7(27,28,29) Status(30)
    """
    out = ['' for _ in range(32)]
    if len(row_values) < 30:
        row_values = row_values + [''] * (30 - len(row_values))
    out[0] = ''  # Source Sheet filled later
    out[1] = row_values[1]   # Discipline
    out[2] = row_values[0]   # #
    out[3] = row_values[2]   # Type
    out[4] = row_values[3]   # Reference
    out[5] = row_values[4]   # Description
    # REV.0..REV.7 (3 cols each) = positions 6..29 in INPUT -> positions 6..29 in output
    for i in range(24):
        out[6 + i] = row_values[5 + i]
    # INPUT has no "Consultant" column; its col 30 (index 29) is "Status" -> goes to MOH/Status
    out[30] = ''              # Consultant (empty for INPUT)
    out[31] = row_values[29]  # MOH/Status
    return out

def map_discipline_row(row_values):
    """Map a discipline sheet row (CIV/EL/PL/HVAC/FF/ELVE — 30 cols) to unified 32-col output.
    Discipline layout: #(1) type(2) Reference(3) Description(4)
                       REV.0(5,6,7) REV.1(8,9,10) ... REV.7(26,27,28) Consultant(29) moh(30)
    """
    out = ['' for _ in range(32)]
    if len(row_values) < 30:
        row_values = row_values + [''] * (30 - len(row_values))
    out[0] = ''  # Source Sheet filled later
    out[1] = ''  # Discipline (filled by caller from sheet name)
    out[2] = row_values[0]   # #
    out[3] = row_values[1]   # type
    out[4] = row_values[2]   # Reference
    out[5] = row_values[3]   # Description
    # REV.0..REV.7 (3 cols each) = positions 4..27 in source (0-indexed) -> positions 6..29 in output
    for i in range(24):
        out[6 + i] = row_values[4 + i]
    out[30] = row_values[28]  # Consultant
    out[31] = row_values[29]  # moh
    return out

def map_29col_row(row_values):
    """Map a 29-col sheet row (CHECK LIST, MOH, soor) to unified 32-col output.
    29-col layout: #(1) type(2) Reference(3) Description(4)
                   REV.0(5,6,7) REV.1(8,9,10) ... REV.7(26,27,28) Status(29)
    (No Consultant/MOH split — only a single Status col)
    """
    out = ['' for _ in range(32)]
    if len(row_values) < 29:
        row_values = row_values + [''] * (29 - len(row_values))
    out[0] = ''  # Source Sheet filled later
    out[1] = ''  # Discipline
    out[2] = row_values[0]   # #
    out[3] = row_values[1]   # type
    out[4] = row_values[2]   # Reference
    out[5] = row_values[3]   # Description
    # REV.0..REV.7 = positions 4..27 in source (0-indexed) -> positions 6..29 in output
    for i in range(24):
        out[6 + i] = row_values[4 + i]
    out[30] = ''              # Consultant (empty)
    out[31] = row_values[28]  # MOH/Status
    return out

def map_cont_row(row_values):
    """Map CONT sheet row (29 cols, but starts at REV.1 — no REV.0).
    CONT layout: #(1) type(2) Reference(3) Description(4)
                 REV.1(5,6,7) REV.2(8,9,10) ... REV.7(26,27,28) Status(29)
    So REV.0 (output cols 6,7,8) will be empty; REV.1..REV.7 shifted accordingly.
    """
    out = ['' for _ in range(32)]
    if len(row_values) < 29:
        row_values = row_values + [''] * (29 - len(row_values))
    out[0] = ''  # Source Sheet filled later
    out[1] = ''  # Discipline
    out[2] = row_values[0]   # #
    out[3] = row_values[1]   # type
    out[4] = row_values[2]   # Reference
    out[5] = row_values[3]   # Description
    # REV.1..REV.7 = positions 4..27 in source (0-indexed) -> positions 9..32 in output
    # Wait, that's wrong. Output positions for REV.1 start at index 9 (REV.1 Submit).
    # And REV.7 Action ends at output index 29 (since output has 32 cols, indices 0..31).
    # REV.1: out indices 9,10,11 ← src indices 4,5,6
    # REV.2: out indices 12,13,14 ← src indices 7,8,9
    # ...
    # REV.7: out indices 27,28,29 ← src indices 25,26,27
    for rev_idx in range(1, 8):
        for sub in range(3):
            src_idx = 4 + (rev_idx - 1) * 3 + sub
            out_idx = 6 + rev_idx * 3 + sub
            if src_idx < len(row_values) and out_idx < 32:
                out[out_idx] = row_values[src_idx]
    out[30] = ''              # Consultant (empty)
    out[31] = row_values[28]  # MOH/Status
    return out

# ---------- Define which sheets to process and how ----------
SHEET_SPECS = [
    # (sheet_name, data_start_row, mapper)
    ('INPUT',      6, map_input_row),
    ('CIV',        6, map_discipline_row),
    ('EL',         6, map_discipline_row),
    ('PL',         6, map_discipline_row),
    ('HVAC',       6, map_discipline_row),
    ('FF',         6, map_discipline_row),
    ('ELVE',       6, map_discipline_row),
    ('CHECK LIST', 6, map_29col_row),
    ('MOH',        6, map_29col_row),
    ('soor',       6, map_29col_row),
    ('CONT',       6, map_cont_row),
]

# ---------- Extract rows ----------
all_rows = []  # list of 32-col lists
stats = {}

for sn, start_row, mapper in SHEET_SPECS:
    if sn not in wb_src.sheetnames:
        print(f'  WARNING: sheet {sn!r} not found, skipping')
        continue
    ws = wb_src[sn]
    sheet_rows = 0
    for r in range(start_row, ws.max_row + 1):
        # Read raw row values (cleaned)
        max_col = ws.max_column
        raw = [clean(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        # Map to unified 32-col output
        mapped = mapper(raw)
        # Set Source Sheet and Discipline
        mapped[0] = sn
        if sn == 'INPUT':
            # Discipline already set from col 2 of INPUT
            pass
        else:
            mapped[1] = discipline_from_sheet(sn)
        all_rows.append(mapped)
        sheet_rows += 1
    stats[sn] = sheet_rows
    print(f'  Extracted {sheet_rows:4d} rows from {sn!r}')

print(f'\nTotal rows extracted: {len(all_rows)}')

# ---------- Sort by Reference ----------
# Reference is in col index 4. Use a sort key that handles empty/numeric parts.
def ref_sort_key(row):
    ref = row[4]
    if not ref:
        return (1, '')  # empty refs go last
    # Try to extract numeric portion for natural sorting
    m = re.match(r'^([A-Za-z\-]+)\s*-?\s*(\d+)', ref)
    if m:
        prefix = m.group(1).upper()
        num = int(m.group(2))
        return (0, prefix, num)
    return (0, ref.upper(), 0)

print('Sorting by Reference...')
all_rows.sort(key=ref_sort_key)

# ---------- Write output ----------
print(f'\nWriting output: {DST}')
os.makedirs(os.path.dirname(DST), exist_ok=True)

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = 'Combined Register'

# Headers
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(border_style='thin', color='B0B0B0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, h in enumerate(UNIFIED_HEADERS, 1):
    cell = ws_out.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Data rows
data_font = Font(name='Calibri', size=10)
data_align_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
data_align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

for i, row in enumerate(all_rows, 2):
    for c, v in enumerate(row, 1):
        cell = ws_out.cell(row=i, column=c, value=v)
        cell.font = data_font
        cell.border = border
        # Center the small columns: Source Sheet, Discipline, #, Type, Reference,
        # all REV Submit/Date Reply/Action, Consultant, MOH/Status
        if c in (1, 2, 3, 4, 5) or c >= 7:
            cell.alignment = data_align_center
        else:
            # Description (col 6) left-aligned, wrapped
            cell.alignment = data_align_top

# Column widths
col_widths = {
    1: 14,   # Source Sheet
    2: 10,   # Discipline
    3: 6,    # #
    4: 16,   # Type
    5: 14,   # Reference
    6: 45,   # Description
}
for col_idx in range(7, 32):
    col_widths[col_idx] = 14  # REV cols and Status cols

for col_idx, w in col_widths.items():
    ws_out.column_dimensions[get_column_letter(col_idx)].width = w

# Freeze header row AND the first 6 identification columns
ws_out.freeze_panes = 'G2'

# Auto-filter on the header row
last_col_letter = get_column_letter(len(UNIFIED_HEADERS))
ws_out.auto_filter.ref = f'A1:{last_col_letter}{len(all_rows) + 1}'

# Workbook metadata
wb_out.properties.creator = 'Z.ai'
wb_out.properties.title = 'LOG Combined — All Sheets'

wb_out.save(DST)

print(f'\nDone. Output saved to: {DST}')
print(f'\nStats per sheet:')
for k, v in stats.items():
    print(f'  {k:14s}: {v:4d} rows')
print(f'\nTotal combined rows: {len(all_rows)}')
print(f'Total columns: {len(UNIFIED_HEADERS)}')
