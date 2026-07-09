"""
Generate a horizontal timeline Excel report from JSON data.

Usage:
    python3 gen_timeline_report.py <input_json> <output_xlsx>

Each transmittal = one row, each revision (REV.0..REV.7) = column group
(Submit Date / Reply Date / Action / Days Open).
"""
import openpyxl
import json
import sys
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IN = sys.argv[1]
OUT = sys.argv[2]

def fmt_date(s):
    if not s:
        return ''
    try:
        dt = datetime.fromisoformat(s.replace('Z', '+00:00'))
        return dt.strftime('%d/%m/%Y')
    except Exception:
        try:
            dt = datetime.strptime(s[:10], '%Y-%m-%d')
            return dt.strftime('%d/%m/%Y')
        except Exception:
            return s

def fmt_action(a):
    if not a:
        return ''
    a = a.lower().strip()
    if a == 'approved': return 'معتمد'
    if a == 'rejected': return 'مرفوض'
    if a == 'withdrawn': return 'مسحوب'
    if a == 'pending': return 'بانتظار'
    return a

def fmt_approval_letter(at):
    """Convert approvalType code to letter A-E."""
    mapping = {
        'APPROVED': 'A',
        'APPROVED_AS_NOTED': 'B',
        'APPROVED_AS_NOTED_RESUBMIT': 'C',
        'NOT_APPROVED': 'D',
        'FOR_INFORMATION': 'E',
    }
    return mapping.get(at, '')

def fmt_approval_label(at):
    """Convert approvalType code to Arabic description."""
    mapping = {
        'APPROVED': 'معتمد',
        'APPROVED_AS_NOTED': 'معتمد بملاحظات',
        'APPROVED_AS_NOTED_RESUBMIT': 'معتمد بملاحظات وإعادة',
        'NOT_APPROVED': 'غير معتمد',
        'FOR_INFORMATION': 'للمعلومات',
    }
    return mapping.get(at, '')

def get_action_fill(action, approval_type=None):
    """Get fill color based on action + approvalType."""
    if not action:
        return None
    a = action.lower().strip()
    if a == 'approved':
        if approval_type == 'NOT_APPROVED':
            return PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # red
        if approval_type in ('FOR_INFORMATION', 'APPROVED_AS_NOTED_RESUBMIT'):
            return PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # yellow/orange
        return PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # green
    if a == 'rejected':
        return PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # red
    if a == 'withdrawn':
        return PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')  # gray
    if a == 'pending':
        return PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # yellow
    return None

with open(IN, 'r', encoding='utf-8') as f:
    items = json.load(f)

print(f'Loaded {len(items)} items', flush=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Timeline Report'
ws.sheet_view.showGridLines = False
ws.sheet_view.rightToLeft = True

# Styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
group_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
group_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
data_font = Font(name='Calibri', size=10)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
right = Alignment(horizontal='right', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin = Side(border_style='thin', color='B0B0B0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Action colors
action_fills = {
    'approved': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'rejected': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'withdrawn': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
    'pending': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
}

# Compute the maximum rev number dynamically from the data.
# This makes the table grow automatically when more revs are added.
max_rev = 0
for item in items:
    for k in (item.get('revisions') or {}).keys():
        try:
            n = int(k)
            if n > max_rev:
                max_rev = n
        except (ValueError, TypeError):
            pass
NUM_REVS = max_rev + 1  # number of rev columns (REV.0..REV.{max_rev})
print(f'Max rev number found in data: {max_rev} → showing {NUM_REVS} columns (REV.0..REV.{max_rev})', flush=True)

# Title row
ws.merge_cells('A1:N1')
ws.cell(row=1, column=1).value = f'تقرير الجدول الزمني للترانسميتالات - Timeline Report ({len(items)} عنصر · REV.0 - REV.{max_rev})'
ws.cell(row=1, column=1).font = Font(name='Calibri', size=14, bold=True, color='1F4E78')
ws.cell(row=1, column=1).alignment = center
ws.row_dimensions[1].height = 30

# Group header row (row 2)
fixed_cols = [
    ('A', 'المرجع'),
    ('B', 'القسم الرئيسي'),
    ('C', 'التخصص'),
    ('D', 'النوع'),
    ('E', 'الوصف'),
]
group_headers = fixed_cols + [
    (None, 'REV.0'), (None, 'REV.1'), (None, 'REV.2'), (None, 'REV.3'),
    (None, 'REV.4'), (None, 'REV.5'), (None, 'REV.6'), (None, 'REV.7'),
    (None, 'الاستشاري'), (None, 'الوزارة'), (None, 'إجمالي الأيام'),
]

# Compute column positions
# A,B,C,D,E = fixed (1-5)
# REV.0 group = 3 cols (6,7,8) = Submit/Reply/Action
# REV.1 = 9,10,11
# ...
# REV.7 = 27,28,29
# Then Consultant(30), MOH(31,32), TotalDays(33)

# (max_rev and NUM_REVS were computed above — reused here)

col_idx = 1
fixed_cols_count = 5
col_idx = fixed_cols_count + 1
rev_start_cols = {}
for r in range(NUM_REVS):
    rev_start_cols[r] = col_idx
    col_idx += 4
consultant_col = col_idx
moh_submit_col = col_idx + 1
moh_reply_col = col_idx + 2
moh_status_col = col_idx + 3
total_days_col = col_idx + 4
last_col = total_days_col

# Group headers row 2
for i, (col_letter, label) in enumerate(fixed_cols, 1):
    ws.cell(row=2, column=i).value = label
    ws.cell(row=2, column=i).font = group_font
    ws.cell(row=2, column=i).fill = group_fill
    ws.cell(row=2, column=i).alignment = center
    ws.cell(row=2, column=i).border = border

# Merge REV groups in row 2 (group header)
for r in range(NUM_REVS):
    start = rev_start_cols[r]
    end = start + 3
    ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
    cell = ws.cell(row=2, column=start)
    cell.value = f'REV.{r}'
    cell.font = group_font
    cell.fill = group_fill
    cell.alignment = center
    cell.border = border
    for c in range(start, end + 1):
        ws.cell(row=2, column=c).border = border
        ws.cell(row=2, column=c).fill = group_fill

# Consultant / MOH / Total headers (row 2)
ws.merge_cells(start_row=2, start_column=consultant_col, end_row=2, end_column=consultant_col)
ws.cell(row=2, column=consultant_col).value = 'الاستشاري'
ws.cell(row=2, column=consultant_col).font = group_font
ws.cell(row=2, column=consultant_col).fill = group_fill
ws.cell(row=2, column=consultant_col).alignment = center
ws.cell(row=2, column=consultant_col).border = border

ws.merge_cells(start_row=2, start_column=moh_submit_col, end_row=2, end_column=moh_status_col)
ws.cell(row=2, column=moh_submit_col).value = 'الوزارة'
ws.cell(row=2, column=moh_submit_col).font = group_font
ws.cell(row=2, column=moh_submit_col).fill = group_fill
ws.cell(row=2, column=moh_submit_col).alignment = center
for c in range(moh_submit_col, moh_status_col + 1):
    ws.cell(row=2, column=c).border = border
    ws.cell(row=2, column=c).fill = group_fill

ws.cell(row=2, column=total_days_col).value = 'إجمالي الأيام'
ws.cell(row=2, column=total_days_col).font = group_font
ws.cell(row=2, column=total_days_col).fill = group_fill
ws.cell(row=2, column=total_days_col).alignment = center
ws.cell(row=2, column=total_days_col).border = border

# Sub-header row (row 3) for each REV group: Submit / Reply / Action
sub_labels = ['تقديم', 'رد', 'إجراء', 'نوع']
for r in range(NUM_REVS):
    start = rev_start_cols[r]
    for i, lbl in enumerate(sub_labels):
        c = ws.cell(row=3, column=start + i)
        c.value = lbl
        c.font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
        c.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        c.alignment = center
        c.border = border

# Sub-header for Consultant (row 3)
ws.cell(row=3, column=consultant_col).value = 'الحالة'
ws.cell(row=3, column=consultant_col).font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
ws.cell(row=3, column=consultant_col).fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ws.cell(row=3, column=consultant_col).alignment = center
ws.cell(row=3, column=consultant_col).border = border

# Sub-headers for MOH
moh_labels = ['إرسال', 'رد', 'الحالة']
for i, lbl in enumerate(moh_labels):
    c = ws.cell(row=3, column=moh_submit_col + i)
    c.value = lbl
    c.font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
    c.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    c.alignment = center
    c.border = border

# Sub-header for Total Days
ws.cell(row=3, column=total_days_col).value = 'يوم'
ws.cell(row=3, column=total_days_col).font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
ws.cell(row=3, column=total_days_col).fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ws.cell(row=3, column=total_days_col).alignment = center
ws.cell(row=3, column=total_days_col).border = border

# Fixed cols also need sub-header (row 3) - leave empty or merged
for i in range(1, fixed_cols_count + 1):
    ws.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)
    ws.cell(row=3, column=i).border = border

# Data rows (starting from row 4)
for idx, item in enumerate(items, 4):
    # Fixed cols
    ws.cell(row=idx, column=1).value = item['reference']
    ws.cell(row=idx, column=2).value = item.get('category', '')
    ws.cell(row=idx, column=3).value = item.get('discipline', '')
    ws.cell(row=idx, column=4).value = item.get('type', '')
    ws.cell(row=idx, column=5).value = item.get('description', '')

    # REV columns
    for r in range(NUM_REVS):
        rev_data = item.get('revisions', {}).get(str(r))
        start = rev_start_cols[r]
        if rev_data:
            ws.cell(row=idx, column=start).value = fmt_date(rev_data.get('submitDate'))
            ws.cell(row=idx, column=start + 1).value = fmt_date(rev_data.get('replyDate'))
            action = rev_data.get('action') or ''
            approval_type = rev_data.get('approvalType') or ''
            ws.cell(row=idx, column=start + 2).value = fmt_action(action)
            # Color the action cell based on action + approvalType
            fill = get_action_fill(action, approval_type)
            if fill:
                ws.cell(row=idx, column=start + 2).fill = fill
            # Approval type letter (A/B/C/D/E) in the 4th sub-column
            letter = fmt_approval_letter(approval_type) if action == 'approved' else ''
            ws.cell(row=idx, column=start + 3).value = letter

    # Consultant
    ws.cell(row=idx, column=consultant_col).value = item.get('consultantStatus', '')

    # MOH
    ws.cell(row=idx, column=moh_submit_col).value = fmt_date(item.get('mohSubmitDate'))
    ws.cell(row=idx, column=moh_reply_col).value = fmt_date(item.get('mohReviewDate'))
    ws.cell(row=idx, column=moh_status_col).value = item.get('mohStatus', '')

    # Total days
    ws.cell(row=idx, column=total_days_col).value = item.get('totalDays', 0)

    # Apply formatting to all cells in this row
    for c in range(1, last_col + 1):
        cell = ws.cell(row=idx, column=c)
        cell.font = data_font
        cell.border = border
        if c == 1:  # Reference - center, bold
            cell.alignment = center
            cell.font = Font(name='Calibri', size=10, bold=True)
        elif c == 5:  # Description - left
            cell.alignment = left
        else:
            cell.alignment = center

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 40
for r in range(NUM_REVS):
    start = rev_start_cols[r]
    for offset, w in enumerate([12, 12, 10, 6]):
        ws.column_dimensions[get_column_letter(start + offset)].width = w
ws.column_dimensions[get_column_letter(consultant_col)].width = 14
ws.column_dimensions[get_column_letter(moh_submit_col)].width = 12
ws.column_dimensions[get_column_letter(moh_reply_col)].width = 12
ws.column_dimensions[get_column_letter(moh_status_col)].width = 14
ws.column_dimensions[get_column_letter(total_days_col)].width = 10

# Freeze panes (freeze first 5 cols + header rows)
ws.freeze_panes = ws.cell(row=4, column=6)

# Save
wb.save(OUT)
print(f'Report saved: {OUT}', flush=True)
print(f'Size: {os.path.getsize(OUT)} bytes', flush=True)
print(f'Rows: {len(items)} transmittals', flush=True)
print(f'Cols: {last_col} (5 fixed + {NUM_REVS}×3 revisions + 4 moh/total)', flush=True)
