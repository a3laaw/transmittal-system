"""Verify the combined output by sampling rows."""
import openpyxl
wb = openpyxl.load_workbook('/home/z/my-project/download/LOG_Combined_AllSheets.xlsx', data_only=True)
ws = wb.active

print(f'Sheet: {ws.title}')
print(f'Total rows (incl. header): {ws.max_row}')
print(f'Total columns: {ws.max_column}')
print()
print('Header:')
for c in range(1, ws.max_column + 1):
    print(f'  C{c:02d} ({openpyxl.utils.get_column_letter(c)}): {ws.cell(row=1, column=c).value}')

print()
print('=' * 130)
print('First 5 data rows (after sort by Reference):')
print('=' * 130)
for r in range(2, 7):
    print(f'\n--- Row {r} ---')
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        v = ws.cell(row=r, column=c).value or ''
        if v:
            print(f'  {h:20s}: {v}')

print()
print('=' * 130)
print('Sample rows with actual data from different source sheets:')
print('=' * 130)
# Find rows from each source sheet (with non-empty Reference)
seen_sheets = set()
shown = 0
for r in range(2, ws.max_row + 1):
    src = ws.cell(row=r, column=1).value
    ref = ws.cell(row=r, column=5).value
    if src and ref and src not in seen_sheets:
        seen_sheets.add(src)
        shown += 1
        print(f'\n--- Row {r} (Source: {src}) ---')
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            v = ws.cell(row=r, column=c).value or ''
            if v:
                print(f'  {h:20s}: {v}')
        if shown >= 11:
            break

print()
print('=' * 130)
print('Summary: rows per source sheet')
print('=' * 130)
from collections import Counter
src_counter = Counter()
for r in range(2, ws.max_row + 1):
    src = ws.cell(row=r, column=1).value or '(empty)'
    src_counter[src] += 1
for k, v in src_counter.most_common():
    print(f'  {k:14s}: {v:4d} rows')
