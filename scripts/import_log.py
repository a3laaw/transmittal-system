"""
Import LOG_Final.xlsm into JSON for the Next.js API to ingest.

Usage:
    python3 import_log.py <input_xlsm> <output_json>

Output JSON format: list of {
    reference, discipline, type, description,
    revisions: [{revNumber, submitDate, replyDate, action}],
    consultantStatus, mohStatus
}

Reads only the 6 discipline sheets: CIV, EL, PL, HVAC, FF, ELVE
Skips rows with no Description.
"""
import openpyxl
import json
import sys
import re
from datetime import datetime

SRC = sys.argv[1]
OUT = sys.argv[2]

DISCIPLINE_SHEETS = ['CIV', 'EL', 'PL', 'HVAC', 'FF', 'ELVE']

def clean(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    s = re.sub(r'\s*\n\s*', ' / ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    # Remove non-breaking spaces
    s = s.replace('\xa0', ' ').strip()
    return s

def parse_action(v):
    """Parse action from Excel cell. Returns (action, approval_type).
    Only 3 actions: approved, rejected, withdrawn.
    Old 'pending' is mapped to '' (no action yet)."""
    s = clean(v).lower()
    if not s:
        return ('', '')
    if 'approved' in s and 'noted' in s and 'resubmit' in s:
        return ('approved', 'APPROVED_AS_NOTED_RESUBMIT')
    if 'approved' in s and 'noted' in s:
        return ('approved', 'APPROVED_AS_NOTED')
    if 'not approved' in s or 'not_approve' in s:
        return ('approved', 'NOT_APPROVED')
    if 'for information' in s or 'more info' in s:
        return ('approved', 'FOR_INFORMATION')
    if s == 'approved' or s == 'approve' or 'approved' in s:
        return ('approved', 'APPROVED')
    if 'reject' in s:
        return ('rejected', '')
    if 'withdrawn' in s or 'withdraw' in s:
        return ('withdrawn', '')
    if 'pending' in s:
        return ('', '')  # no action yet - don't store 'pending'
    return (s, '')

def main():
    print(f'Loading: {SRC}', flush=True)
    wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
    print(f'Sheets: {wb.sheetnames}', flush=True)

    out = []
    for sn in DISCIPLINE_SHEETS:
        if sn not in wb.sheetnames:
            print(f'  WARNING: {sn} not found', flush=True)
            continue
        ws = wb[sn]
        # Header is at rows 4-5, data starts at row 6
        # Cols: #(1) type(2) Reference(3) Description(4)
        # REV.0: Submit(5) DateReply(6) Action(7)
        # REV.1: Submit(8) DateReply(9) Action(10) ... 3 cols each
        # REV.7: Submit(26) DateReply(27) Action(28)
        # Consultant(29) moh(30)
        rows_in_sheet = 0
        for r in range(6, ws.max_row + 1):
            reference = clean(ws.cell(row=r, column=3).value)
            description = clean(ws.cell(row=r, column=4).value)
            type_val = clean(ws.cell(row=r, column=2).value)

            # Skip empty rows (no description AND no reference)
            if not description and not reference:
                continue
            # Skip rows with no description (per user request)
            if not description:
                continue

            # Build revisions array
            revisions = []
            for rev_idx in range(8):  # REV.0..REV.7
                submit = clean(ws.cell(row=r, column=5 + rev_idx * 3).value)
                reply = clean(ws.cell(row=r, column=6 + rev_idx * 3).value)
                action, approval_type = parse_action(ws.cell(row=r, column=7 + rev_idx * 3).value)
                # Only include this rev if at least submit or action exists
                if submit or action:
                    revisions.append({
                        'revNumber': rev_idx,
                        'submitDate': submit,
                        'replyDate': reply,
                        'action': action,
                        'approvalType': approval_type if action == 'approved' else '',
                    })

            consultant = clean(ws.cell(row=r, column=29).value)
            moh = clean(ws.cell(row=r, column=30).value)

            # Normalize review statuses
            def norm_status(s):
                if not s: return ''
                s_lower = s.lower()
                if 'approved' in s_lower and 'overdue' not in s_lower:
                    return 'Approved'
                if 'overdue' in s_lower:
                    return 'Overdue'
                if 'under review' in s_lower or 'pending' in s_lower:
                    return 'Under Review'
                if 'cancelled' in s_lower:
                    return 'Cancelled'
                if 'submit' in s_lower and 'rev' in s_lower:
                    return 'Submit Next Rev'
                return s

            out.append({
                'reference': reference,
                'discipline': sn,
                'type': type_val,
                'description': description,
                'revisions': revisions,
                'consultantStatus': norm_status(consultant),
                'mohStatus': norm_status(moh),
            })
            rows_in_sheet += 1
        print(f'  {sn}: {rows_in_sheet} rows', flush=True)

    # Write JSON
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f'\nTotal extracted: {len(out)}', flush=True)
    print(f'Written to: {OUT}', flush=True)

if __name__ == '__main__':
    main()
