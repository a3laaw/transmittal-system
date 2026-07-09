"""
Create a clean, fully-formatted Excel template from TRANSIMITALS.xlsx sheet '170'.

The template preserves ALL borders, fills, fonts, merged cells, column widths,
and row heights from the original. Only cell VALUES in the item area (rows 16-26)
are cleared — formatting stays intact.
"""
import openpyxl
import os
import shutil

SRC = '/home/z/my-project/upload/TRANSIMITALS.xlsx'
TEMPLATE_DIR = '/home/z/my-project/public/templates'
TEMPLATE = f'{TEMPLATE_DIR}/TRANSIMITALS_template.xlsx'

os.makedirs(TEMPLATE_DIR, exist_ok=True)

print(f'Loading source: {SRC}')
wb = openpyxl.load_workbook(SRC)

# Use sheet '170' as base (has clean structure with 1-2 items)
BASE_SHEET = '170'
if BASE_SHEET not in wb.sheetnames:
    BASE_SHEET = wb.sheetnames[0]
print(f'Base sheet: {BASE_SHEET}')

# Remove all other sheets
for sn in list(wb.sheetnames):
    if sn != BASE_SHEET:
        del wb[sn]

ws = wb[BASE_SHEET]

# Clear ONLY values in item rows (16-26), keep all borders/formatting
cleared = 0
for r in range(16, 27):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        # Skip merged cells (only top-left is writable)
        if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
            if cell.value is not None:
                try:
                    cell.value = None
                    cleared += 1
                except (AttributeError, TypeError):
                    pass

print(f'Cleared {cleared} cell values in item area (rows 16-26), kept all borders/formatting')

# Reset header cells to placeholders
ws.cell(row=3, column=7).value = 'Transmittal No:'  # G3
ws.cell(row=3, column=9).value = 'Rev.00'           # I3
ws.cell(row=4, column=7).value = 'Date :'           # G4

# Rename sheet
ws.title = 'TEMPLATE'

# Make sure sheet view shows the table area clearly
ws.sheet_view.showGridLines = False  # Keep gridlines off (matches original)
ws.sheet_view.zoomScale = 100

# Save
wb.save(TEMPLATE)
print(f'Template saved: {TEMPLATE}')
print(f'Size: {os.path.getsize(TEMPLATE)} bytes')

# Verify borders are preserved
print()
print('Verification - borders in item area:')
wb2 = openpyxl.load_workbook(TEMPLATE)
ws2 = wb2.active
bordered_rows = []
for r in range(14, 28):
    has_border = False
    for c in range(2, 9):
        cell = ws2.cell(row=r, column=c)
        b = cell.border
        if (b.left and b.left.style) or (b.right and b.right.style) or (b.top and b.top.style) or (b.bottom and b.bottom.style):
            has_border = True
            break
    if has_border:
        bordered_rows.append(r)
print(f'  Bordered rows: {bordered_rows}')
print(f'  Total bordered rows in table area: {len(bordered_rows)}')

# Verify column widths
print()
print('Column widths:')
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    cd = ws2.column_dimensions.get(col_letter)
    print(f'  {col_letter}: width={cd.width if cd else "default"}')

# Verify row heights
print()
print('Row heights:')
for r in range(1, 35):
    rd = ws2.row_dimensions.get(r)
    if rd and rd.height:
        print(f'  R{r}: height={rd.height}')

# Verify merged cells
print()
print(f'Merged cell ranges: {len(ws2.merged_cells.ranges)}')
for rng in list(ws2.merged_cells.ranges)[:10]:
    print(f'  {rng}')
