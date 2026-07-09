import openpyxl

wb = openpyxl.load_workbook('/home/z/my-project/download/TRANSMITTALS_Extracted_Rev01plus.xlsx', data_only=True)
ws = wb.active

print(f'Sheet: {ws.title}')
print(f'Total rows (incl. header): {ws.max_row}')
print(f'Total columns: {ws.max_column}')
print()
print('Full content:')
print('=' * 130)
header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print(' | '.join(str(h) for h in header))
print('-' * 130)
for r in range(2, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value or '' for c in range(1, ws.max_column + 1)]
    # Truncate DESCRIPTION for readability
    desc = str(row[4])
    if len(desc) > 60:
        desc = desc[:57] + '...'
    row[4] = desc
    print(' | '.join(str(v) for v in row))

# Summary by sheet
print()
print('=' * 130)
print('Summary by sheet:')
from collections import Counter
sheet_counts = Counter()
for r in range(2, ws.max_row + 1):
    sn = ws.cell(row=r, column=1).value
    sheet_counts[sn] += 1
for sn, cnt in sheet_counts.items():
    tn = ws.cell(row=2 + sum(sheet_counts[k] for k in list(sheet_counts.keys()) if k < sn), column=2).value if False else ''
    # Get first occurrence
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == sn:
            tn = ws.cell(row=r, column=2).value
            rev = ws.cell(row=r, column=3).value
            dt = ws.cell(row=r, column=4).value
            break
    print(f'  {sn:15s} | {tn:12s} | {rev:8s} | {dt:12s} | {cnt} item(s)')
