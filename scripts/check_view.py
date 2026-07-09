"""Check sheet view settings, gridlines, and other display options."""
import openpyxl

print("=" * 80)
print("Sheet view settings comparison")
print("=" * 80)

# Original
wb1 = openpyxl.load_workbook('/home/z/my-project/upload/TRANSIMITALS.xlsx')
ws1 = wb1['170']
print("\nORIGINAL '170':")
print(f'  sheet_view.showGridLines: {ws1.sheet_view.showGridLines}')
print(f'  sheet_view.showRowColHeaders: {ws1.sheet_view.showRowColHeaders}')
print(f'  sheet_view.showOutlineSymbols: {ws1.sheet_view.showOutlineSymbols}')
print(f'  sheet_view.defaultGridColor: {ws1.sheet_view.defaultGridColor}')

# Template
wb2 = openpyxl.load_workbook('/home/z/my-project/public/templates/TRANSIMITALS_template.xlsx')
ws2 = wb2.active
print("\nCURRENT TEMPLATE:")
print(f'  sheet_view.showGridLines: {ws2.sheet_view.showGridLines}')
print(f'  sheet_view.showRowColHeaders: {ws2.sheet_view.showRowColHeaders}')
print(f'  sheet_view.showOutlineSymbols: {ws2.sheet_view.showOutlineSymbols}')

# Check what the generated file looks like
print()
print("=" * 80)
print("Generated file (test_out.xlsx from earlier)")
print("=" * 80)
wb3 = openpyxl.load_workbook('/tmp/test_out.xlsx')
ws3 = wb3.active
print(f'  Sheet: {ws3.title}')
print(f'  showGridLines: {ws3.sheet_view.showGridLines}')

# Count rows with borders in template (item area, rows 16-27)
print()
print("=" * 80)
print("Template: count of bordered rows in item area (rows 16-27)")
print("=" * 80)
bordered_rows = []
for r in range(16, 28):
    has_border = False
    for c in range(2, 9):
        cell = ws2.cell(row=r, column=c)
        b = cell.border
        if (b.left and b.left.style) or (b.right and b.right.style) or (b.top and b.top.style) or (b.bottom and b.bottom.style):
            has_border = True
            break
    if has_border:
        bordered_rows.append(r)
print(f'  Rows with borders: {bordered_rows}')

# Compare with original
bordered_rows_orig = []
for r in range(16, 28):
    has_border = False
    for c in range(2, 9):
        cell = ws1.cell(row=r, column=c)
        b = cell.border
        if (b.left and b.left.style) or (b.right and b.right.style) or (b.top and b.top.style) or (b.bottom and b.bottom.style):
            has_border = True
            break
    if has_border:
        bordered_rows_orig.append(r)
print(f'  Original rows with borders: {bordered_rows_orig}')
