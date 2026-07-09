import openpyxl
wb = openpyxl.load_workbook('/home/z/my-project/upload/TRANSIMITALS.xlsx', data_only=True)

# Look at a few sheets to understand structure
for sn in ['146 (2)', '170', '27', '27-D01' if '27-D01' in wb.sheetnames else '27', '1', '1r1']:
    if sn not in wb.sheetnames:
        continue
    print('=' * 80)
    print('SHEET:', sn)
    print('=' * 80)
    ws = wb[sn]
    for r in range(1, min(ws.max_row + 1, 35)):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row_vals.append('')
            else:
                row_vals.append(str(v))
        print(f"R{r:02d}: " + ' | '.join(row_vals))
    print()
