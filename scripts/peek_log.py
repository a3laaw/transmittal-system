import openpyxl

# Use data_only=False to see formulas, but we'll also try data_only=True to see calculated values
wb = openpyxl.load_workbook('/home/z/my-project/upload/LOG_Final.xlsm', data_only=True, keep_vba=True)

print('Sheet names:', wb.sheetnames)
print()

# Look at each sheet briefly - first 20 rows
for sn in wb.sheetnames:
    print('=' * 100)
    print(f'SHEET: {sn}')
    print('=' * 100)
    ws = wb[sn]
    print(f'Dimensions: {ws.dimensions} | max_row={ws.max_row} | max_col={ws.max_column}')
    print()
    # Print first 15 rows, non-empty columns only
    for r in range(1, min(ws.max_row + 1, 16)):
        row_vals = []
        for c in range(1, min(ws.max_column + 1, 31)):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row_vals.append('')
            else:
                s = str(v).replace('\n', ' / ')
                if len(s) > 25:
                    s = s[:22] + '...'
                row_vals.append(s)
        # Compact print
        line = f'R{r:02d}: '
        for i, v in enumerate(row_vals):
            if v:
                line += f'C{i+1}={v!r} | '
        print(line.rstrip(' | '))
    print()
